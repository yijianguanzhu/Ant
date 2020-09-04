# !/usr/bin/python
# coding=utf-8

import openpyxl
from openpyxl.styles import Font


# 定义一个写入excel .xlsx 的类
class WriteExcelXlsx:

    def __init__(self, path: str = None, sheet_name: str = None):
        # excel文件位置路径
        self.path = path
        self.sheet_name = sheet_name
        self.wk = openpyxl.load_workbook(path)

    def write_excel(self, int_row: int, int_col: int, content: str):
        sheet = self.wk[self.sheet_name]
        # 填充内容
        sheet.cell(int_row, int_col, content)
        # 设置字体，字号
        font = Font(name='Calibri', size=9)
        sheet.cell(int_row, int_col).font = font
